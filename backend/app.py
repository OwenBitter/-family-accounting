"""Flask API for family accounting application."""

import json
import os
import tempfile
import threading
import time
from pathlib import Path

import openpyxl

from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS

import config
from models.schemas import MonthlyData, AssetSnapshot, IncomeRecord
from services.data_store import DataStore
from services.excel_reader import parse_alipay_csv, parse_wechat_xlsx
from services.excel_writer import create_monthly_book
from services.classifier import classify, classify_income, get_all_categories

app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": ["http://localhost:5173",
                                               "http://localhost:5174",
                                               "http://localhost:3000",
                                               "http://127.0.0.1:5173",
                                               "http://127.0.0.1:5174"]}})

app.config["MAX_CONTENT_LENGTH"] = config.MAX_CONTENT_LENGTH
app.config["UPLOAD_FOLDER"] = str(config.UPLOAD_DIR)
app.config["DATA_FOLDER"] = str(config.DATA_DIR)

config.UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
config.HISTORY_DIR.mkdir(parents=True, exist_ok=True)

data_store = DataStore(config.HISTORY_DIR)

# Background cleanup thread for old temp files
def _cleanup_temp_files():
    while True:
        time.sleep(3600)  # every hour
        now = time.time()
        for f in config.UPLOAD_DIR.glob("*"):
            if now - f.stat().st_mtime > 3600:
                f.unlink(missing_ok=True)

cleanup_thread = threading.Thread(target=_cleanup_temp_files, daemon=True)
cleanup_thread.start()


# ─── Error handling ────────────────────────────────────────────────

class AppError(Exception):
    def __init__(self, code: str, message: str, status: int = 400, details=None):
        self.code = code
        self.message = message
        self.status = status
        self.details = details or {}


@app.errorhandler(AppError)
def handle_app_error(error):
    return jsonify({
        "success": False,
        "error": {
            "code": error.code,
            "message": error.message,
            "details": error.details,
        }
    }), error.status


@app.errorhandler(413)
def handle_too_large(error):
    return jsonify({
        "success": False,
        "error": {
            "code": "FILE_TOO_LARGE",
            "message": "文件超过大小限制（最大16MB）",
            "details": {},
        }
    }), 413


# ─── Helpers ───────────────────────────────────────────────────────

def _save_upload(file_storage) -> str:
    """Save uploaded file to temp dir and return path."""
    ext = Path(file_storage.filename).suffix.lower() if file_storage.filename else ".tmp"
    if ext not in config.ALLOWED_EXTENSIONS:
        raise AppError("INVALID_FILE_FORMAT", f"不支持的文件格式: {ext}", details={"ext": ext})
    fd, path = tempfile.mkstemp(suffix=ext, dir=str(config.UPLOAD_DIR))
    os.close(fd)
    file_storage.save(path)
    return path


def _parse_files(file_paths, person, auto_detect):
    """Parse uploaded files and return transactions."""
    all_transactions = []
    person_warnings = []

    for fp in file_paths:
        ext = Path(fp).suffix.lower()
        try:
            if ext == ".csv":
                txns = parse_alipay_csv(fp, person=person, auto_detect=auto_detect)
            else:
                txns = parse_wechat_xlsx(fp, person=person, auto_detect=auto_detect)
        except Exception as e:
            raise AppError("PARSE_ERROR", f"文件解析失败: {e}", details={"file": fp})

        for t in txns:
            if t.amount > 0:
                t.target_category = classify_income(t.description, t.counterparty)
            else:
                t.target_category = classify(t.source, t.raw_category,
                                              t.description, t.amount)

        # Person mismatch detection
        if auto_detect and txns:
            detected = txns[0].detected_person
            if detected and detected != person:
                person_warnings.append({
                    "file": Path(fp).name,
                    "detected_person": detected,
                })

        all_transactions.extend(txns)

        # Clean up temp file
        try:
            os.unlink(fp)
        except OSError:
            pass

    return all_transactions, person_warnings


# ─── Health ─────────────────────────────────────────────────────────

@app.route("/api/health")
def health():
    return jsonify({"status": "ok"})


# ─── Import ─────────────────────────────────────────────────────────

@app.route("/api/import/preview", methods=["POST"])
def import_preview():
    """Upload files and return parsed + classified transactions."""
    person = request.form.get("person", "")
    month = request.form.get("month", "")
    auto_detect = request.form.get("auto_detect", "true").lower() == "true"

    if person not in config.PERSONS:
        raise AppError("INVALID_PARAM", "请选择人员（BB/LN）", details={"person": person})

    files = request.files.getlist("files")
    if not files:
        raise AppError("INVALID_PARAM", "请至少上传一个文件")

    file_paths = [_save_upload(f) for f in files]
    transactions, person_warnings = _parse_files(file_paths, person, auto_detect)

    # Find duplicates if existing data
    existing = data_store.get_month(month)
    duplicate_ids = []
    if existing:
        existing_ids = {(t.order_id, t.amount)
                        for t in existing.expenses.get(person, [])}
        for t in transactions:
            if (t.order_id, t.amount) in existing_ids:
                duplicate_ids.append(t.order_id)

    # Statistics
    total_expense = sum(abs(t.amount) for t in transactions if t.amount < 0)
    total_income = sum(t.amount for t in transactions if t.amount > 0)
    unmapped = [t for t in transactions if t.target_category == "其他" and t.amount < 0]

    return jsonify({
        "success": True,
        "transactions": [t.to_dict() for t in transactions],
        "statistics": {
            "total_count": len(transactions),
            "expense_count": sum(1 for t in transactions if t.amount < 0),
            "income_count": sum(1 for t in transactions if t.amount > 0),
            "total_expense": round(total_expense, 2),
            "total_income": round(total_income, 2),
            "unmapped_count": len(unmapped),
        },
        "duplicate_order_ids": duplicate_ids,
        "person_warnings": person_warnings,
    })


@app.route("/api/import/confirm", methods=["POST"])
def import_confirm():
    """Confirm import, write to Excel and cache."""
    data = request.get_json()
    if not data:
        raise AppError("INVALID_PARAM", "请求体不能为空")

    person = data.get("person", "")
    month = data.get("month", "")
    transactions_data = data.get("transactions", [])

    if person not in config.PERSONS or not month:
        raise AppError("INVALID_PARAM", "人员或月份参数无效")

    # Load or create monthly data
    monthly = data_store.get_month(month)
    if monthly is None:
        monthly = MonthlyData(month=month)

    # Parse and add transactions (separate income vs expense)
    from models.schemas import Transaction, IncomeRecord
    from datetime import datetime
    added_count = 0
    for td in transactions_data:
        t = Transaction.from_dict(td)
        if t.amount > 0:
            income = IncomeRecord(
                person=person,
                time=t.time,
                category=t.target_category or "其他",
                amount=t.amount,
                channel="支付宝" if t.source == "alipay" else "微信",
                account=t.counterparty,
                note=t.description,
            )
            monthly.income[person].append(income)
        else:
            monthly.expenses[person].append(t)
        added_count += 1

    # Write to Excel
    try:
        file_path = create_monthly_book(month, monthly)
    except Exception as e:
        raise AppError("EXCEL_WRITE_ERROR", f"Excel 写入失败: {e}")

    # Save to cache
    data_store.save_month(month, monthly)
    data_store.rebuild_index()

    return jsonify({
        "success": True,
        "filePath": file_path,
        "added_count": added_count,
    })


# ─── Data ───────────────────────────────────────────────────────────

def _prev_month(month: str) -> str:
    """Return previous month string, e.g. '2026.5' → '2026.4'."""
    parts = month.split(".")
    if len(parts) != 2:
        return ""
    y, m = int(parts[0]), int(parts[1])
    if m == 1:
        return f"{y - 1}.12"
    return f"{y}.{m - 1}"


@app.route("/api/data/summary")
def data_summary():
    """Get monthly summary (stats for dashboard stat cards)."""
    month = request.args.get("month", "")
    monthly = data_store.get_month(month)

    if not monthly:
        return jsonify({"success": True, "data": None})

    bb_income = sum(r.amount for r in monthly.income.get("BB", []))
    ln_income = sum(r.amount for r in monthly.income.get("LN", []))
    # Fallback to analysis data if no individual transactions
    bb_exp = monthly.expenses.get("BB", [])
    ln_exp = monthly.expenses.get("LN", [])
    bb_expense = sum(abs(t.amount) for t in bb_exp if t.amount < 0) if bb_exp else sum(monthly.analysis.get("BB", {}).values())
    ln_expense = sum(abs(t.amount) for t in ln_exp if t.amount < 0) if ln_exp else sum(monthly.analysis.get("LN", {}).values())
    bb_asset = monthly.assets.get("BB")
    ln_asset = monthly.assets.get("LN")

    bb_total = round(bb_asset.total, 2) if bb_asset else 0
    ln_total = round(ln_asset.total, 2) if ln_asset else 0
    external_asset = monthly.external_asset or 0
    other_asset = monthly.other_asset or 0
    # grandTotal = 个人资产 + 外借资产（不重复加其他资产，因可能是总计数）
    grand_total = round(bb_total + ln_total + external_asset, 2)

    # Compute "本月攒" as asset change: current month minus previous month
    prev_monthly = data_store.get_month(_prev_month(month))
    if prev_monthly and prev_monthly.assets.get("BB") and prev_monthly.assets.get("LN"):
        prev_bb = prev_monthly.assets["BB"].total
        prev_ln = prev_monthly.assets["LN"].total
        bb_saved = round(bb_total - prev_bb, 2) if bb_asset else 0
        ln_saved = round(ln_total - prev_ln, 2) if ln_asset else 0
    else:
        # Fallback to income - expense if no previous month data
        bb_saved = round(bb_income - bb_expense, 2)
        ln_saved = round(ln_income - ln_expense, 2)

    result = {
        "month": month,
        "bb": {
            "lastBalance": monthly.last_balance.get("BB", 0),
            "income": round(bb_income, 2),
            "expense": round(bb_expense, 2),
            "saved": bb_saved,
            "total": bb_total,
        },
        "ln": {
            "lastBalance": monthly.last_balance.get("LN", 0),
            "income": round(ln_income, 2),
            "expense": round(ln_expense, 2),
            "saved": ln_saved,
            "total": ln_total,
        },
        "total": {
            "income": round(bb_income + ln_income, 2),
            "expense": round(bb_expense + ln_expense, 2),
            "saved": round(bb_saved + ln_saved, 2),
            "grandTotal": grand_total,
            "externalAsset": round(external_asset, 2),
            "otherAsset": round(other_asset, 2),
        },
    }

    return jsonify({"success": True, "data": result})


@app.route("/api/data/expenses")
def data_expenses():
    """Get expense details and category analysis for a month."""
    month = request.args.get("month", "")
    monthly = data_store.get_month(month)

    if not monthly:
        return jsonify({"success": True, "details": [], "analysis": []})

    details = []
    analysis_map: dict[str, dict] = {}

    has_transactions = any(len(monthly.expenses.get(p, [])) > 0 for p in ["BB", "LN"])

    if has_transactions:
        for person in ["BB", "LN"]:
            for t in monthly.expenses.get(person, []):
                if t.amount >= 0:
                    continue
                d = t.to_dict()
                cat = t.target_category or classify(t.source, t.raw_category, t.description, t.amount, t.counterparty)
                d["liveCategory"] = cat
                details.append(d)
                if cat not in analysis_map:
                    analysis_map[cat] = {"category": cat, "bbAmount": 0, "lnAmount": 0, "totalAmount": 0}
                analysis_map[cat][f"{person.lower()}Amount"] += abs(t.amount)
    else:
        # Fallback: use pre-aggregated analysis from xlsx
        std_cats = get_all_categories()
        def normalize_cat(name: str) -> str:
            """Match xlsx category names like '娱乐（阅读、游戏...）' to standard '娱乐'."""
            for s in std_cats:
                if name == s or name.startswith(s):
                    return s
            return name

        for person, pkey in [("BB", "bbAmount"), ("LN", "lnAmount")]:
            for cat, amt in monthly.analysis.get(person, {}).items():
                norm = normalize_cat(cat)
                if norm not in analysis_map:
                    analysis_map[norm] = {"category": norm, "bbAmount": 0, "lnAmount": 0, "totalAmount": 0}
                analysis_map[norm][pkey] = round(amt, 2)

    for cat in get_all_categories():
        if cat not in analysis_map:
            analysis_map[cat] = {"category": cat, "bbAmount": 0, "lnAmount": 0, "totalAmount": 0}

    for cat_data in analysis_map.values():
        cat_data["totalAmount"] = round(cat_data["bbAmount"] + cat_data["lnAmount"], 2)
        cat_data["bbAmount"] = round(cat_data["bbAmount"], 2)
        cat_data["lnAmount"] = round(cat_data["lnAmount"], 2)

    analysis = [analysis_map[cat] for cat in get_all_categories()
                if cat in analysis_map]

    return jsonify({
        "success": True,
        "details": details,
        "analysis": analysis,
    })


@app.route("/api/data/assets")
def data_assets():
    """Get asset data for a month."""
    month = request.args.get("month", "")
    monthly = data_store.get_month(month)

    if not monthly:
        return jsonify({"success": True, "data": []})

    result = []
    for person in ["BB", "LN"]:
        asset = monthly.assets.get(person)
        if asset:
            d = asset.to_dict()
            d["total"] = asset.total  # ensure total is always present
            result.append(d)

    return jsonify({"success": True, "data": result})


@app.route("/api/data/assets", methods=["PUT"])
def update_assets():
    """Update asset data for a person in a month."""
    data = request.get_json()
    if not data:
        raise AppError("INVALID_PARAM", "请求体不能为空")

    month = data.get("month", "")
    person = data.get("person", "")

    if person not in config.PERSONS or not month:
        raise AppError("INVALID_PARAM", "人员或月份参数无效")

    monthly = data_store.get_month(month)
    if monthly is None:
        monthly = MonthlyData(month=month)

    asset = AssetSnapshot(
        person=person,
        month=month,
        alipay_fund=data.get("alipay_fund", 0),
        alipay_yuebao=data.get("alipay_yuebao", 0),
        alipay_balance=data.get("alipay_balance", 0),
        wechat_balance=data.get("wechat_balance", 0),
        wechat_licaitong=data.get("wechat_licaitong", 0),
        bank_accounts=data.get("bank_accounts", {}),
        other=data.get("other", {}),
        loan_receivable=data.get("loan_receivable", 0),
    )
    monthly.assets[person] = asset
    monthly.last_balance[person] = asset.total

    try:
        file_path = create_monthly_book(month, monthly)
    except Exception as e:
        raise AppError("EXCEL_WRITE_ERROR", f"Excel 写入失败: {e}")

    data_store.save_month(month, monthly)
    data_store.rebuild_index()

    return jsonify({"success": True, "filePath": file_path})


@app.route("/api/data/income")
def data_income():
    """Get income records for a month."""
    month = request.args.get("month", "")
    monthly = data_store.get_month(month)

    if not monthly:
        return jsonify({"success": True, "records": []})

    records = []
    for person in ["BB", "LN"]:
        for r in monthly.income.get(person, []):
            records.append(r.to_dict())

    return jsonify({"success": True, "records": records})


@app.route("/api/data/income", methods=["PUT"])
def update_income():
    """Replace all income records for a person in a month."""
    data = request.get_json()
    if not data:
        raise AppError("INVALID_PARAM", "请求体不能为空")

    month = data.get("month", "")
    person = data.get("person", "")

    if person not in config.PERSONS or not month:
        raise AppError("INVALID_PARAM", "人员或月份参数无效")

    monthly = data_store.get_month(month)
    if monthly is None:
        monthly = MonthlyData(month=month)

    records = data.get("records", [])
    from datetime import datetime
    new_income = []
    for rec in records:
        try:
            t = datetime.strptime(rec.get("time", ""), "%Y-%m-%d") if rec.get("time") else datetime.now()
        except (ValueError, TypeError):
            t = datetime.now()
        new_income.append(IncomeRecord(
            person=person,
            time=t,
            category=rec.get("category", "其他"),
            amount=rec.get("amount", 0),
            channel=rec.get("channel", ""),
            account=rec.get("account", ""),
            note=rec.get("note", ""),
        ))

    monthly.income[person] = new_income

    try:
        file_path = create_monthly_book(month, monthly)
    except Exception as e:
        raise AppError("EXCEL_WRITE_ERROR", f"Excel 写入失败: {e}")

    data_store.save_month(month, monthly)
    data_store.rebuild_index()

    return jsonify({"success": True, "filePath": file_path})


@app.route("/api/data/history")
def data_history():
    """Get list of available months."""
    months = data_store.get_history()
    return jsonify({"success": True, "availableMonths": months})


@app.route("/api/data/trend")
def data_trend():
    """Get trend data for all available months (for dashboard charts)."""
    months = data_store.get_history()
    if not months:
        return jsonify({"success": True, "data": []})

    trend_data = []
    prev_monthly = None
    for month in months:
        monthly = data_store.get_month(month)
        if not monthly:
            continue

        bb_exp = monthly.expenses.get("BB", [])
        ln_exp = monthly.expenses.get("LN", [])
        bb_expense = sum(abs(t.amount) for t in bb_exp if t.amount < 0) if bb_exp else sum(monthly.analysis.get("BB", {}).values())
        ln_expense = sum(abs(t.amount) for t in ln_exp if t.amount < 0) if ln_exp else sum(monthly.analysis.get("LN", {}).values())
        bb_income = sum(r.amount for r in monthly.income.get("BB", []))
        ln_income = sum(r.amount for r in monthly.income.get("LN", []))

        # "本月攒" as asset change: current minus previous month
        bb_asset = monthly.assets.get("BB")
        ln_asset = monthly.assets.get("LN")
        bb_total = round(bb_asset.total, 2) if bb_asset else 0
        ln_total = round(ln_asset.total, 2) if ln_asset else 0

        if prev_monthly and prev_monthly.assets.get("BB") and prev_monthly.assets.get("LN"):
            prev_bb = prev_monthly.assets["BB"].total
            prev_ln = prev_monthly.assets["LN"].total
            saved = round((bb_total - prev_bb) + (ln_total - prev_ln), 2)
        else:
            saved = round((bb_income + ln_income) - (bb_expense + ln_expense), 2)

        trend_data.append({
            "month": month,
            "income": round(bb_income + ln_income, 2),
            "expense": round(bb_expense + ln_expense, 2),
            "saved": saved,
        })

        prev_monthly = monthly

    return jsonify({"success": True, "data": trend_data})


# ─── Investments (loan book + gold) ────────────────────────────────

INVESTMENT_FILE = config.INVESTMENT_FILE
GOLD_FILE = config.GOLD_FILE


@app.route("/api/data/investments")
def data_investments():
    """Read loan book and gold inventory from Excel."""
    result = {"loanBook": [], "gold": []}

    if INVESTMENT_FILE.exists():
        wb = openpyxl.load_workbook(INVESTMENT_FILE, data_only=True)
        if "进账" in wb.sheetnames:
            ws = wb["进账"]
            current_month = ""
            for r in range(3, ws.max_row + 1):
                month_v = ws.cell(r, 1).value
                if month_v:
                    current_month = str(month_v).strip()
                person = ws.cell(r, 2).value
                amount = ws.cell(r, 3).value
                if not person or not amount:
                    continue
                result["loanBook"].append({
                    "month": current_month,
                    "person": str(person).strip(),
                    "amount": float(amount),
                    "note": str(ws.cell(r, 4).value or "").strip(),
                    "totalInCard": float(ws.cell(r, 5).value) if ws.cell(r, 5).value else None,
                })
        wb.close()

    if GOLD_FILE.exists():
        wb = openpyxl.load_workbook(GOLD_FILE, data_only=True)
        ws = wb.active
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, 1).value
            weight = ws.cell(r, 3).value
            source = ws.cell(r, 4).value
            if not name:
                continue
            result["gold"].append({
                "name": str(name).strip(),
                "weight": float(weight) if weight else 0,
                "source": str(source or "").strip(),
                "imageIndex": r - 1,  # 1-based index for image file
            })
        wb.close()

    return jsonify({"success": True, **result})


# ─── Gold price ────────────────────────────────────────────────────

_gold_price_cache = {"price": 0, "time": 0}

@app.route("/api/data/gold-price")
def gold_price():
    """Fetch current gold price per gram in CNY, cached for 1 hour."""
    import time as ttime
    now = ttime.time()
    if now - _gold_price_cache["time"] < 3600 and _gold_price_cache["price"] > 0:
        return jsonify({"success": True, "pricePerGram": _gold_price_cache["price"]})

    try:
        import urllib.request, json, ssl
        ctx = ssl.create_default_context()
        # Fetch gold price in USD/oz
        req = urllib.request.Request(
            "https://api.gold-api.com/price/XAU",
            headers={"User-Agent": "Mozilla/5.0"},
        )
        resp = urllib.request.urlopen(req, context=ctx, timeout=10)
        data = json.loads(resp.read().decode())
        usd_per_oz = float(data["price"])

        # Fetch USD/CNY exchange rate
        req2 = urllib.request.Request(
            "https://api.exchangerate-api.com/v4/latest/USD",
            headers={"User-Agent": "Mozilla/5.0"},
        )
        resp2 = urllib.request.urlopen(req2, context=ctx, timeout=10)
        fx = json.loads(resp2.read().decode())
        cny_per_usd = float(fx["rates"]["CNY"])

        # Convert: USD/oz → CNY/g
        oz_to_g = 31.1035
        price_per_g = round(usd_per_oz / oz_to_g * cny_per_usd, 2)

        _gold_price_cache["price"] = price_per_g
        _gold_price_cache["time"] = now
        return jsonify({"success": True, "pricePerGram": price_per_g})
    except Exception as e:
        return jsonify({"success": False, "pricePerGram": 0, "error": str(e)})


# ─── Gold image ────────────────────────────────────────────────────

@app.route("/api/data/gold-image/<int:idx>")
def gold_image(idx):
    """Serve gold item image from xlsx by index (1-based)."""
    import zipfile
    if not GOLD_FILE.exists():
        return jsonify({"error": "file not found"}), 404
    try:
        with zipfile.ZipFile(GOLD_FILE) as z:
            img_path = f"xl/media/image{idx}.png"
            if img_path not in z.namelist():
                return jsonify({"error": "image not found"}), 404
            img_data = z.read(img_path)
        from flask import Response
        return Response(img_data, mimetype="image/png")
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Categories ────────────────────────────────────────────────────

@app.route("/api/categories")
def list_categories():
    """Get all categories and current mappings."""
    return jsonify({
        "success": True,
        "categories": get_all_categories(),
        "mappings": config.CATEGORY_MAP,
    })


# ─── Export ─────────────────────────────────────────────────────────

@app.route("/api/export")
def export_month():
    """Download Excel file for a given month."""
    month = request.args.get("month", "")
    if not month:
        raise AppError("INVALID_PARAM", "请指定月份")

    file_path = config.HISTORY_DIR / f"家庭收支{month}.xlsx"
    if not file_path.exists():
        raise AppError("FILE_NOT_FOUND", f"{month} 的记账文件不存在", status=404)

    return send_file(
        str(file_path),
        as_attachment=True,
        download_name=f"家庭收支{month}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ─── Frontend static files ──────────────────────────────────────────

FRONTEND_DIST = config.BASE_DIR.parent / "frontend" / "dist"


@app.route("/")
@app.route("/<path:path>")
def serve_frontend(path=""):
    """Serve frontend static files. SPA fallback to index.html."""
    # Don't intercept API routes (Flask route priority handles this,
    # but being explicit avoids confusion)
    if path.startswith("api/"):
        from flask import abort
        abort(404)

    if path and (FRONTEND_DIST / path).exists():
        return send_from_directory(str(FRONTEND_DIST), path)
    return send_from_directory(str(FRONTEND_DIST), "index.html")


# ─── Main ──────────────────────────────────────────────────────────

if __name__ == "__main__":
    print(f"Starting server on http://localhost:5000")
    print(f"Data directory: {config.DATA_DIR}")
    app.run(host="0.0.0.0", port=5000, debug=True)
