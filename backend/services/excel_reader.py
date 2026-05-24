"""Parse Alipay CSV and WeChat xlsx files into standardized Transaction records."""

import csv
import io
import re
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl

from config import ALIPAY_ACCOUNTS, WECHAT_NICKNAMES
from models.schemas import Transaction


def parse_alipay_csv(file_path: str, person: str = "",
                     auto_detect: bool = True) -> list[Transaction]:
    """Parse Alipay CSV file (GBK encoded, with metadata header)."""
    transactions: list[Transaction] = []
    detected_person: Optional[str] = None

    with open(file_path, encoding="gbk", errors="replace") as f:
        content = f.read()

    # Auto-detect person from account info in metadata
    if auto_detect:
        detected_person = _detect_alipay_person(content)

    lines = content.splitlines()
    if not lines:
        return transactions

    # Find header row (contains column names)
    header_idx = _find_alipay_header(lines)
    if header_idx < 0:
        return transactions

    headers = _parse_csv_line(lines[header_idx])
    col_map = _build_column_map(headers, {
        "交易时间": "time",
        "交易分类": "raw_category",
        "交易对方": "counterparty",
        "对方账号": "counterparty_account",
        "商品说明": "description",
        "收/支": "type",
        "金额": "amount",
        "收/付款方式": "payment_method",
        "交易状态": "status",
        "交易订单号": "order_id",
        "商家订单号": "merchant_order_id",
        "备注": "memo",
    })

    person_confidence = "auto_detected" if detected_person else "confirmed"
    effective_person = detected_person if detected_person else person

    for i in range(header_idx + 1, len(lines)):
        line = lines[i].strip()
        if not line:
            continue
        row = _parse_csv_line(line)
        if len(row) < len(col_map):
            continue

        t = _build_transaction(row, col_map, "alipay", effective_person or person,
                                detected_person, person_confidence)
        if t is not None:
            transactions.append(t)

    return transactions


def parse_wechat_xlsx(file_path: str, person: str = "",
                      auto_detect: bool = True) -> list[Transaction]:
    """Parse WeChat xlsx file (with metadata header rows)."""
    transactions: list[Transaction] = []
    detected_person: Optional[str] = None

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return transactions

    # Auto-detect person from nickname in metadata
    if auto_detect:
        metadata_text = " ".join([str(r[0]) for r in rows[:20] if r[0]])
        detected_person = _detect_wechat_person(metadata_text)

    # Find header row
    header_idx = _find_wechat_header(rows)
    if header_idx < 0:
        return transactions

    headers = list(rows[header_idx])
    col_map = _build_column_map(headers, {
        "交易时间": "time",
        "交易类型": "raw_category",
        "交易对方": "counterparty",
        "商品": "description",
        "收/支": "type",
        "金额(元)": "amount",
        "支付方式": "payment_method",
        "当前状态": "status",
        "交易单号": "order_id",
        "商户单号": "merchant_order_id",
        "备注": "memo",
    })

    person_confidence = "auto_detected" if detected_person else "confirmed"
    effective_person = detected_person if detected_person else person

    for i in range(header_idx + 1, len(rows)):
        row = list(rows[i])
        if not any(row):
            continue
        if len(row) < len(col_map):
            continue

        t = _build_transaction(row, col_map, "wechat", effective_person or person,
                                detected_person, person_confidence)
        if t is not None:
            transactions.append(t)

    return transactions


def read_existing_book(path: str, month: str) -> dict:
    """Read existing xlsx to extract stored data for the given month."""
    wb = openpyxl.load_workbook(path, data_only=True)
    result = {
        "month": month,
        "income": {"BB": [], "LN": []},
        "expenses": {"BB": [], "LN": []},
        "analysis": {},
        "assets": {},
    }

    # Read 收入 sheet
    if "收入" in wb.sheetnames:
        ws = wb["收入"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            person_label = str(row[0] or "")
            person = "BB" if person_label == "斌" else "LN" if person_label == "纳" else ""
            if person and row[1]:
                result["income"][person].append({
                    "time": str(row[1]),
                    "category": str(row[2] or ""),
                    "amount": float(row[3] or 0),
                    "channel": str(row[4] or ""),
                    "account": str(row[5] or ""),
                    "note": str(row[6] or ""),
                })

    # Read 支出明细 sheet
    if "支出明细" in wb.sheetnames:
        ws = wb["支出明细"]
        current_person = ""
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            p = str(row[0] or "")
            if p in ("斌", "纳"):
                current_person = "BB" if p == "斌" else "LN"
            if current_person and row[1]:
                result["expenses"][current_person].append({
                    "time": str(row[1]),
                    "raw_category": str(row[2] or ""),
                    "amount": abs(float(row[3] or 0)),
                    "payment_method": str(row[4] or ""),
                    "description": str(row[5] or ""),
                })

    # Read 支出分析 sheet
    if "支出分析" in wb.sheetnames:
        ws = wb["支出分析"]
        for row in ws.iter_rows(min_row=3, max_row=13, values_only=True):
            if row[0]:
                result["analysis"][str(row[0])] = {
                    "LN": float(row[1] or 0),
                    "BB": float(row[2] or 0),
                    "total": float(row[3] or 0),
                }

    wb.close()
    return result


def _find_alipay_header(lines: list[str]) -> int:
    for i, line in enumerate(lines):
        if "交易时间" in line and "交易分类" in line:
            return i
    return -1


def _find_wechat_header(rows: list[tuple]) -> int:
    for i, row in enumerate(rows):
        if row and "交易时间" in str(row[0]):
            return i
    return -1


def _parse_csv_line(line: str) -> list[str]:
    """Parse a CSV line handling quoted fields."""
    reader = csv.reader(io.StringIO(line))
    return next(reader, [])


def _build_column_map(headers: list, expected: dict) -> dict:
    """Build a mapping from column index to our field name."""
    col_map = {}
    for i, h in enumerate(headers):
        h_str = str(h).strip() if h else ""
        if h_str in expected:
            col_map[i] = expected[h_str]
    return col_map


def _build_transaction(row, col_map, source, person, detected_person, person_confidence):
    """Convert a parsed row to a Transaction object."""
    fields = {}
    for col_idx, field_name in col_map.items():
        val = row[col_idx] if col_idx < len(row) else ""
        fields[field_name] = val

    # Parse time
    time_str = str(fields.get("time", "")).strip()
    time = _parse_datetime(time_str)
    if time is None:
        return None

    # Parse type (收入/支出)
    txn_type = str(fields.get("type", "")).strip()

    # Parse amount
    amount_str = str(fields.get("amount", "")).strip().replace(",", "").replace("¥", "").replace("￥", "")
    try:
        amount = float(amount_str)
    except (ValueError, TypeError):
        amount = 0.0

    if txn_type == "支出":
        amount = -abs(amount)
    elif txn_type == "收入":
        amount = abs(amount)

    raw_category = str(fields.get("raw_category", "")).strip()

    return Transaction(
        person=person,
        source=source,
        time=time,
        raw_category=raw_category,
        target_category="",
        amount=amount,
        payment_method=str(fields.get("payment_method", "")).strip(),
        description=str(fields.get("description", fields.get("memo", ""))).strip(),
        counterparty=str(fields.get("counterparty", "")).strip(),
        status=str(fields.get("status", "")).strip(),
        order_id=str(fields.get("order_id", "")).strip(),
        detected_person=detected_person,
        person_confidence=person_confidence,
    )


def _parse_datetime(s: str) -> Optional[datetime]:
    """Try multiple datetime formats."""
    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt)
        except (ValueError, TypeError):
            continue
    return None


def _detect_alipay_person(content: str) -> Optional[str]:
    """Detect person from Alipay CSV metadata."""
    content_lower = content.lower()
    for person, info in ALIPAY_ACCOUNTS.items():
        for kw in info["keywords"]:
            if kw.lower() in content_lower:
                return person
    return None


def _detect_wechat_person(metadata: str) -> Optional[str]:
    """Detect person from WeChat xlsx metadata."""
    metadata_lower = metadata.lower()
    for person, info in WECHAT_NICKNAMES.items():
        for kw in info["keywords"]:
            if kw.lower() in metadata_lower:
                return person
    return None
