"""Write standardized Excel files matching the 家庭收支 template format."""

import shutil
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def _safe_set_cell(ws, row, column, value=None):
    """Set cell value, gracefully handling MergedCell."""
    # Check if cell is part of a merged range
    for merge_range in list(ws.merged_cells.ranges):
        if (merge_range.min_row <= row <= merge_range.max_row and
            merge_range.min_col <= column <= merge_range.max_col):
            ws.unmerge_cells(str(merge_range))
            cell = ws.cell(row=row, column=column)
            cell.value = value
            return cell
    cell = ws.cell(row=row, column=column)
    cell.value = value
    return cell

from config import DATA_DIR, TEMPLATE_FILE, CATEGORIES, PERSON_LABELS
from models.schemas import MonthlyData


def create_monthly_book(month: str, data: MonthlyData,
                        output_dir: Optional[Path] = None) -> str:
    """Create or update a monthly Excel file.

    Args:
        month: Month string like "2026.1"
        data: MonthlyData with all transactions, incomes, assets
        output_dir: Output directory (default: data/history/)

    Returns:
        Path to the generated file
    """
    if output_dir is None:
        output_dir = DATA_DIR / "history"
    output_dir.mkdir(parents=True, exist_ok=True)

    output_path = output_dir / f"家庭收支{month}.xlsx"
    month_num = month.split(".")[-1]

    # Load template or existing file
    if output_path.exists():
        wb = openpyxl.load_workbook(output_path)
        _clear_data_sheets(wb)
    elif TEMPLATE_FILE.exists():
        shutil.copy2(TEMPLATE_FILE, output_path)
        wb = openpyxl.load_workbook(output_path)
    else:
        raise FileNotFoundError(f"Template not found: {TEMPLATE_FILE}")

    _write_summary_sheet(wb, month_num, data)
    _write_income_sheet(wb, data)
    _write_expenses_sheet(wb, data)
    _write_analysis_sheet(wb, data)
    _write_assets_sheet(wb, data)

    wb.save(output_path)
    wb.close()
    return str(output_path)


def _clear_data_sheets(wb):
    """Clear data from all sheets while preserving formulas and structure."""
    clear_configs = {
        "收入": {"start": 2},
        "支出明细": {"start": 2},
        "支出分析": {"start": 3, "end": 13},
    }
    for sheet_name, cfg in clear_configs.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=cfg["start"],
                                max_row=cfg.get("end", ws.max_row)):
            for cell in row:
                if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    cell.value = None


def _write_summary_sheet(wb, month_num: str, data: MonthlyData):
    """Write 总 (summary) sheet."""
    ws = wb["总"]
    ws["D2"] = f"{month_num}月"

    for person, row_num in [("BB", 4), ("LN", 5)]:
        _safe_set_cell(ws, row_num, 3, value=data.last_balance.get(person, 0.0))
        ws.cell(row=row_num, column=3).number_format = '#,##0.00'

        total_income = sum(r.amount for r in data.income.get(person, []))
        _safe_set_cell(ws, row_num, 4, value=total_income)
        ws.cell(row=row_num, column=4).number_format = '#,##0.00'

        total_expense = sum(
            abs(t.amount) for t in data.expenses.get(person, []) if t.amount < 0
        )
        _safe_set_cell(ws, row_num, 5, value=total_expense)
        ws.cell(row=row_num, column=5).number_format = '#,##0.00'

        asset = data.assets.get(person)
        if asset:
            _safe_set_cell(ws, row_num, 7, value=asset.total)
            ws.cell(row=row_num, column=7).number_format = '#,##0.00'

    _safe_set_cell(ws, 7, 7, value=data.external_asset)
    ws.cell(row=7, column=7).number_format = '#,##0.00'
    _safe_set_cell(ws, 8, 7, value=data.other_asset)
    ws.cell(row=8, column=7).number_format = '#,##0.00'


def _write_income_sheet(wb, data: MonthlyData):
    """Write 收入 (income) sheet."""
    ws = wb["收入"]
    row = 2

    thin = Side(style="thin", color="999999")
    border = Border(top=thin, bottom=thin)

    for person in ["BB", "LN"]:
        label = PERSON_LABELS[person]
        for rec in data.income.get(person, []):
            _safe_set_cell(ws, row, 1, value=label)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")

            _safe_set_cell(ws, row, 2,
                           value=rec.time if isinstance(rec.time, str)
                           else rec.time.strftime("%Y-%m-%d %H:%M:%S"))

            _safe_set_cell(ws, row, 3, value=rec.category)
            _safe_set_cell(ws, row, 4, value=rec.amount)
            ws.cell(row=row, column=4).number_format = '#,##0.00'

            _safe_set_cell(ws, row, 5, value=rec.channel)
            _safe_set_cell(ws, row, 6, value=rec.account)
            _safe_set_cell(ws, row, 7, value=rec.note)

            row += 1


def _write_expenses_sheet(wb, data: MonthlyData):
    """Write 支出明细 (expense details) sheet."""
    ws = wb["支出明细"]
    row = 2

    thin = Side(style="thin", color="999999")
    border = Border(top=thin, bottom=thin)

    for person in ["BB", "LN"]:
        label = PERSON_LABELS[person]
        txns = data.expenses.get(person, [])
        if not txns:
            continue

        start_row = row
        for t in txns:
            if t.amount >= 0:
                continue  # Only expenses

            _safe_set_cell(ws, row, 1, value=label)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")

            _safe_set_cell(ws, row, 2,
                           value=t.time.strftime("%Y-%m-%d %H:%M:%S"))

            _safe_set_cell(ws, row, 3, value=t.target_category or t.raw_category)
            _safe_set_cell(ws, row, 4, value=abs(t.amount))
            ws.cell(row=row, column=4).number_format = '#,##0.00'

            _safe_set_cell(ws, row, 5, value=t.payment_method)
            _safe_set_cell(ws, row, 6, value=t.description)

            row += 1

        end_row = row - 1
        if end_row >= start_row:
            try:
                ws.merge_cells(start_row=start_row, start_column=1,
                               end_row=end_row, end_column=1)
            except openpyxl.utils.exceptions.IllegalCharacterError:
                pass


def _write_analysis_sheet(wb, data: MonthlyData):
    """Write 支出分析 (expense analysis) sheet."""
    ws = wb["支出分析"]

    # Calculate category totals
    bb_totals: dict[str, float] = {}
    ln_totals: dict[str, float] = {}

    for t in data.expenses.get("BB", []):
        if t.amount >= 0:
            continue
        cat = t.target_category or t.raw_category
        bb_totals[cat] = bb_totals.get(cat, 0) + abs(t.amount)

    for t in data.expenses.get("LN", []):
        if t.amount >= 0:
            continue
        cat = t.target_category or t.raw_category
        ln_totals[cat] = ln_totals.get(cat, 0) + abs(t.amount)

    for i, cat in enumerate(CATEGORIES, start=3):
        ln_amt = ln_totals.get(cat, 0)
        bb_amt = bb_totals.get(cat, 0)
        total = ln_amt + bb_amt

        _safe_set_cell(ws, i, 2, value=cat)
        _safe_set_cell(ws, i, 3, value=ln_amt if ln_amt else 0)
        ws.cell(row=i, column=3).number_format = '#,##0.00'
        _safe_set_cell(ws, i, 4, value=bb_amt if bb_amt else 0)
        ws.cell(row=i, column=4).number_format = '#,##0.00'
        _safe_set_cell(ws, i, 5, value=total if total else 0)
        ws.cell(row=i, column=5).number_format = '#,##0.00'


def _write_assets_sheet(wb, data: MonthlyData):
    """Write 理财 (assets) sheet."""
    ws = wb["理财"]

    asset_sections = {
        "BB": {"row": 4, "bank_col_start": 7},
        "LN": {"row": 9, "bank_col_start": 7},
    }

    for person, cfg in asset_sections.items():
        asset = data.assets.get(person)
        if not asset:
            continue

        r = cfg["row"]
        _safe_set_cell(ws, r, 4, value=asset.alipay_fund)
        ws.cell(row=r, column=4).number_format = '#,##0.00'
        _safe_set_cell(ws, r, 5, value=asset.alipay_yuebao)
        ws.cell(row=r, column=5).number_format = '#,##0.00'
        _safe_set_cell(ws, r, 6, value=asset.alipay_balance)
        ws.cell(row=r, column=6).number_format = '#,##0.00'
        _safe_set_cell(ws, r+1, 6, value=asset.wechat_balance)
        ws.cell(row=r+1, column=6).number_format = '#,##0.00'
        _safe_set_cell(ws, r+1, 7, value=asset.wechat_licaitong)
        ws.cell(row=r+1, column=7).number_format = '#,##0.00'

        bank_col = cfg["bank_col_start"]
        for bank_name, amount in asset.bank_accounts.items():
            _safe_set_cell(ws, r, bank_col, value=amount)
            ws.cell(row=r, column=bank_col).number_format = '#,##0.00'
            bank_col += 1

        other_col = bank_col + 1
        for name, amount in asset.other.items():
            _safe_set_cell(ws, r, other_col, value=amount)
            ws.cell(row=r, column=other_col).number_format = '#,##0.00'
            other_col += 1
