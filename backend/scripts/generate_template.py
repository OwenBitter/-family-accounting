"""Generate empty xlsx template from an existing monthly file, clearing data rows."""

import argparse
import shutil
from pathlib import Path
import openpyxl

TEMPLATE_DIR = Path(__file__).resolve().parent.parent.parent / "data" / "templates"
TEMPLATE_FILE = TEMPLATE_DIR / "empty_book.xlsx"

CLEAR_RANGES = {
    "收入": {"start_row": 2},
    "支出明细": {"start_row": 2},
    "支出分析": {"start_row": 3, "end_row": 13},
}

CLEAR_CELLS = {
    "理财": ["D4", "E4", "F4", "G4", "H4", "I4", "J4", "K4", "L4", "M4", "N4",
             "D5", "E5", "F5", "G5", "H5", "I5", "J5", "K5", "L5", "M5", "N5",
             "D9", "E9", "F9", "G9", "H9", "I9", "J9", "K9", "L9", "M9", "N9",
             "D10", "E10", "F10", "G10", "H10", "I10", "J10", "K10", "L10", "M10", "N10",
             "F15"],
    "总": ["D4", "D5", "C4", "C5", "D2"],
}


def clear_cell(ws, cell_ref):
    cell = ws[cell_ref]
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        # Find the master cell of the merge range
        for merge_range in ws.merged_cells.ranges:
            if cell_ref in merge_range:
                master_cell = ws[merge_range.min_row][merge_range.min_col - 1]
                master_cell.value = None
                break
    else:
        cell.value = None


def main():
    parser = argparse.ArgumentParser(description="Generate empty xlsx template from a reference file")
    parser.add_argument("reference", nargs="?", default="",
                        help="Path to an existing monthly xlsx to use as template (default: empty template)")
    parser.add_argument("-o", "--output", default=str(TEMPLATE_FILE),
                        help="Output template path (default: %(default)s)")
    args = parser.parse_args()

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if not args.reference or not Path(args.reference).exists():
        if args.reference:
            print(f"Reference file not found: {args.reference}")
        print("Creating a minimal empty template instead.")
        _create_minimal_template(output_path)
        return

    ref_path = Path(args.reference)
    shutil.copy2(ref_path, output_path)
    wb = openpyxl.load_workbook(output_path)

    for sheet_name, cfg in CLEAR_RANGES.items():
        ws = wb[sheet_name]
        start = cfg["start_row"]
        end = cfg.get("end_row", ws.max_row)
        for row in ws.iter_rows(min_row=start, max_row=end):
            for cell in row:
                if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    cell.value = None

    for sheet_name, cells in CLEAR_CELLS.items():
        ws = wb[sheet_name]
        for cell_ref in cells:
            clear_cell(ws, cell_ref)

    wb.save(output_path)
    print(f"Template created: {output_path}")


def _create_minimal_template(output_path: Path):
    """Create a basic template when no reference file is available."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    # Create required sheets
    for name in ["收入", "支出明细", "支出分析", "理财"]:
        wb.create_sheet(name)
    ws_main = wb["总"] if "总" in wb.sheetnames else wb.active
    ws_main.title = "总"

    sheet_configs = {
        "收入": ["人员", "入账时间", "账务类型", "收入(+元)", "支付渠道", "对方账户", "备注"],
        "支出明细": ["人员", "出账时间", "账务类型", "支出(-元)", "支付渠道", "备注"],
        "支出分析": ["账务类型", "LN支出", "BB支出", "总支出"],
    }
    for sname, headers in sheet_configs.items():
        ws = wb[sname]
        for i, h in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=h)

    wb.save(TEMPLATE_FILE)
    print(f"Minimal template created: {TEMPLATE_FILE}")


if __name__ == "__main__":
    main()
