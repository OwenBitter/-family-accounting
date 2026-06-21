"""Scan billing directory for existing xlsx files and migrate to JSON cache.

Reads all 家庭收支*.xlsx files from year subdirectories,
extracts both individual transactions (支出明细/收入) and summary data.

Usage:
  python scripts/migrate_history.py                    # Use default BILL_DIR
  python scripts/migrate_history.py "E:/path/to/bills" # Custom path
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from models.schemas import MonthlyData, IncomeRecord, Transaction, AssetSnapshot
from services.data_store import DataStore
from services.xlsx_utils import sf, parse_datetime, norm_cat, INCOME_NORM, detect_person_from_label


# Default billing directory. Override via command-line argument.
BILL_DIR = Path("E:/账单")
HISTORY_DIR = Path(__file__).resolve().parent.parent.parent / "data" / "history"


def parse_xlsx_to_monthly(filepath: Path) -> MonthlyData | None:
    month_key = re.sub(r'[\(（][^)）]*[\)）]', '', filepath.stem.replace("家庭收支", "")).strip()
    if not month_key or "总" in month_key or "年" in month_key:
        return None

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"  Cannot open {filepath.name}: {e}")
        return None

    md = MonthlyData(month=month_key)


    # ── 总 sheet ──
    if "总" in wb.sheetnames:
        ws = wb["总"]
        md.last_balance["BB"] = sf(ws.cell(4, 3).value)
        md.last_balance["LN"] = sf(ws.cell(5, 3).value)
        md.external_asset = sf(ws.cell(7, 7).value)
        md.other_asset = sf(ws.cell(8, 7).value)

    # ── 收入 sheet ──
    if "收入" in wb.sheetnames:
        ws = wb["收入"]
        current_person = ""
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(v is not None for v in row):
                continue
            label = str(row[0] or "").strip()
            if "斌" in label:
                current_person = "BB"
            elif "纳" in label or "娜" in label:
                current_person = "LN"
            if not current_person or not row[1] or not row[3]:
                continue
            t = parse_datetime(row[1])
            if not t:
                continue
            cat = INCOME_NORM.get(str(row[2] or "").strip(), "其他")
            md.income[current_person].append(IncomeRecord(
                person=current_person, time=t, category=cat,
                amount=sf(row[3]),
                channel=str(row[4] or "").strip(),
                account=str(row[5] or "").strip(),
                note=str(row[6] or "").strip(),
            ))

    # ── 支出明细 sheet ──
    if "支出明细" in wb.sheetnames:
        ws = wb["支出明细"]
        current_person = ""
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(v is not None for v in row):
                continue
            label = str(row[0] or "").strip()
            if "斌" in label:
                current_person = "BB"
            elif "纳" in label or "娜" in label:
                current_person = "LN"
            if not current_person or not row[1] or not row[3]:
                continue
            t = parse_datetime(row[1])
            if not t:
                continue
            # Handle amount (may be negative in sheet or positive)
            amt = sf(row[3])
            if amt > 0:
                amt = -amt  # expenses are negative
            raw_cat = str(row[2] or "").strip()
            md.expenses[current_person].append(Transaction(
                person=current_person, source="xlsx", time=t,
                raw_category=raw_cat,
                target_category=norm_cat(raw_cat),
                amount=amt,
                payment_method=str(row[4] or "").strip(),
                description=str(row[5] or "").strip(),
                counterparty="", status="",
                order_id="",
                detected_person=current_person,
                person_confidence="confirmed",
            ))

    # ── 支出分析 sheet (for pre-aggregated data as fallback) ──
    if "支出分析" in wb.sheetnames:
        ws = wb["支出分析"]
        for r in range(3, 15):
            cat = ws.cell(r, 2).value
            if cat and str(cat).strip():
                cn = norm_cat(str(cat).strip())
                md.analysis["BB"][cn] = sf(ws.cell(r, 4).value)
                md.analysis["LN"][cn] = sf(ws.cell(r, 3).value)

    # ── 理财 sheet ──
    if "理财" in wb.sheetnames:
        ws = wb["理财"]
        for person, kw in [("BB", "斌"), ("LN", "纳")]:
            asset_row = None
            for r in range(3, 12):
                if kw in str(ws.cell(r, 2).value or ""):
                    asset_row = r
                    break
            if asset_row:
                ast = AssetSnapshot(person=person, month=month_key)
                ast.alipay_fund = sf(ws.cell(asset_row, 4).value)
                ast.alipay_yuebao = sf(ws.cell(asset_row, 5).value)
                ast.alipay_balance = sf(ws.cell(asset_row, 6).value)
                ast.wechat_balance = sf(ws.cell(asset_row + 1, 6).value) if asset_row + 1 <= ws.max_row else 0
                ast.wechat_licaitong = sf(ws.cell(asset_row + 1, 7).value) if asset_row + 1 <= ws.max_row else 0
                for c in range(9, 16):
                    label = ws.cell(asset_row - 1, c).value or ws.cell(asset_row - 2, c).value or ""
                    val = ws.cell(asset_row, c).value
                    if val and str(label).strip():
                        ast.bank_accounts[str(label).strip()] = sf(val)
                md.assets[person] = ast

    wb.close()
    return md


def main():
    HISTORY_DIR.mkdir(parents=True, exist_ok=True)
    store = DataStore(HISTORY_DIR)

    # Clear old cache
    import shutil
    for f in HISTORY_DIR.glob("*.json"):
        if f.name != "index.json":
            f.unlink()

    xlsx_files = sorted(BILL_DIR.rglob("家庭收支*.xlsx"))
    count = 0
    for fp in xlsx_files:
        print(f"Reading {fp.relative_to(BILL_DIR)}...")
        md = parse_xlsx_to_monthly(fp)
        if md is None:
            print(f"  Skip")
            continue
        store.save_month(md.month, md)
        count += 1
        bb_e = len(md.expenses["BB"])
        ln_e = len(md.expenses["LN"])
        bb_i = len(md.income["BB"])
        ln_i = len(md.income["LN"])
        print(f"  OK {md.month}: BB={bb_e}笔支出/{bb_i}笔收入 | LN={ln_e}笔支出/{ln_i}笔收入 | 外借={md.external_asset:.0f}")

    store.rebuild_index()
    print(f"\nDone. Migrated {count} months.")
    print(f"Available: {store.get_history()}")


if __name__ == "__main__":
    main()
