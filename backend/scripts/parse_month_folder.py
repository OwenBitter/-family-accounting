"""Parse a month's folder of xlsx + images for import.

Usage:
  python parse_month_folder.py <folder> [--output manifest.json]
  python parse_month_folder.py <folder> --stdout

Scans folder for:
  - 家庭收支*.xlsx → full monthly data (expenses, income, assets, analysis)
  - 支出*.xlsx / 收入*.xlsx → partial data
  - *.png / *.jpg / *.jpeg → screenshots for Claude to analyze

Outputs a JSON manifest with parsed xlsx data and image listing.
Claude can then read each image (multimodal) and fill in the asset fields.
"""

import json, sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from models.schemas import AssetSnapshot
from services.xlsx_utils import sf, parse_datetime, norm_cat, INCOME_NORM, detect_month_from_path, detect_person_from_label


def parse_single_xlsx(filepath: Path) -> dict:
    """Parse a xlsx and return dict with expenses, income, assets, analysis."""
    result = {
        "expenses": {"BB": [], "LN": []},
        "income": {"BB": [], "LN": []},
        "assets": {"BB": None, "LN": None},
        "analysis": {"BB": {}, "LN": {}},
        "last_balance": {"BB": 0.0, "LN": 0.0},
        "external_asset": 0.0,
        "other_asset": 0.0,
    }

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"  [WARN] Cannot open {filepath.name}: {e}", file=sys.stderr)
        return result

    # ── 总 sheet ──
    if "总" in wb.sheetnames:
        ws = wb["总"]
        result["last_balance"]["BB"] = sf(ws.cell(4, 3).value)
        result["last_balance"]["LN"] = sf(ws.cell(5, 3).value)
        result["external_asset"] = sf(ws.cell(7, 7).value)
        result["other_asset"] = sf(ws.cell(8, 7).value)

    # ── 收入 sheet ──
    if "收入" in wb.sheetnames:
        ws = wb["收入"]
        current_person = ""
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(v is not None for v in row):
                continue
            label = str(row[0] or "").strip()
            if "斌" in label:
                current_person = "BB"
            elif "纳" in label or "娜" in label:
                current_person = "LN"
            if not current_person or not row[1] or not row[3]:
                continue
            t = parse_datetime(row[1])
            if not t:
                continue
            cat = INCOME_NORM.get(str(row[2] or "").strip(), "其他")
            result["income"][current_person].append({
                "person": current_person,
                "time": t.isoformat(),
                "category": cat,
                "amount": sf(row[3]),
                "channel": str(row[4] or "").strip(),
                "account": str(row[5] or "").strip(),
                "note": str(row[6] or "").strip(),
            })

    # ── 支出明细 sheet ──
    if "支出明细" in wb.sheetnames:
        ws = wb["支出明细"]
        current_person = ""
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(v is not None for v in row):
                continue
            label = str(row[0] or "").strip()
            if "斌" in label:
                current_person = "BB"
            elif "纳" in label or "娜" in label:
                current_person = "LN"
            if not current_person or not row[1] or not row[3]:
                continue
            t = parse_datetime(row[1])
            if not t:
                continue
            amt = sf(row[3])
            if amt > 0:
                amt = -amt
            raw_cat = str(row[2] or "").strip()
            result["expenses"][current_person].append({
                "person": current_person, "source": "xlsx",
                "time": t.isoformat(),
                "raw_category": raw_cat,
                "target_category": norm_cat(raw_cat),
                "amount": amt,
                "payment_method": str(row[4] or "").strip(),
                "description": str(row[5] or "").strip(),
                "counterparty": "", "status": "",
                "order_id": "",
                "detected_person": current_person,
                "person_confidence": "confirmed",
            })

    # ── 支出分析 sheet ──
    if "支出分析" in wb.sheetnames:
        ws = wb["支出分析"]
        for r in range(3, 15):
            cat = ws.cell(r, 2).value
            if cat and str(cat).strip():
                cn = norm_cat(str(cat).strip())
                result["analysis"]["BB"][cn] = sf(ws.cell(r, 4).value)
                result["analysis"]["LN"][cn] = sf(ws.cell(r, 3).value)

    # ── 理财 sheet ──
    if "理财" in wb.sheetnames:
        ws = wb["理财"]
        for person, kw in [("BB", "斌"), ("LN", "纳")]:
            asset_row = None
            for r in range(3, 12):
                if kw in str(ws.cell(r, 2).value or ""):
                    asset_row = r
                    break
            if asset_row:
                ast = AssetSnapshot(person=person, month="")
                ast.alipay_fund = sf(ws.cell(asset_row, 4).value)
                ast.alipay_yuebao = sf(ws.cell(asset_row, 5).value)
                ast.alipay_balance = sf(ws.cell(asset_row, 6).value)
                if asset_row + 1 <= ws.max_row:
                    ast.wechat_balance = sf(ws.cell(asset_row + 1, 6).value)
                    ast.wechat_licaitong = sf(ws.cell(asset_row + 1, 7).value)
                for c in range(9, 16):
                    label = ws.cell(asset_row - 1, c).value or ws.cell(asset_row - 2, c).value or ""
                    val = ws.cell(asset_row, c).value
                    if val and str(label).strip():
                        ast.bank_accounts[str(label).strip()] = sf(val)
                result["assets"][person] = ast.to_dict()

    wb.close()
    return result


def scan_folder(folder: Path) -> dict:
    """Scan folder and return manifest."""
    folder = Path(folder).resolve()
    if not folder.exists():
        print(f"Error: folder not found: {folder}", file=sys.stderr)
        sys.exit(1)

    month_key = detect_month_from_path(folder)
    result = {
        "folder": str(folder),
        "month": month_key or "",
        "note": "月从文件夹名自动识别" if month_key else "无法从文件夹名识别月份，请在导入时手动指定",
        "xlsx_data": {
            "expenses": {"BB": [], "LN": []},
            "income": {"BB": [], "LN": []},
            "assets": {"BB": None, "LN": None},
            "analysis": {"BB": {}, "LN": {}},
            "last_balance": {"BB": 0.0, "LN": 0.0},
            "external_asset": 0.0,
            "other_asset": 0.0,
        },
        "images": [],
        "person_dirs": {},
    }

    # Collect xlsx files
    xlsx_files = list(folder.glob("*.xlsx"))

    # Check for person subdirectories (e.g., currMonth/BB/, currMonth/LN/)
    person_dir_map = {}
    for p in ["BB", "LN"]:
        pd = folder / p
        if pd.is_dir():
            files = list(pd.iterdir())
            person_dir_map[p] = [str(f.relative_to(folder)) for f in files]
            xlsx_files.extend(pd.glob("*.xlsx"))

    result["person_dirs"] = person_dir_map

    # Parse xlsx files
    for fp in sorted(set(xlsx_files)):
        print(f"  Parsing {fp.relative_to(folder)}...", file=sys.stderr)
        data = parse_single_xlsx(fp)
        # Merge into result
        for person in ["BB", "LN"]:
            result["xlsx_data"]["expenses"][person].extend(data["expenses"][person])
            result["xlsx_data"]["income"][person].extend(data["income"][person])
            if data["assets"][person]:
                result["xlsx_data"]["assets"][person] = data["assets"][person]
            for cat, amt in data["analysis"][person].items():
                result["xlsx_data"]["analysis"][person][cat] = \
                    result["xlsx_data"]["analysis"][person].get(cat, 0) + amt
        for k in ["last_balance", "external_asset", "other_asset"]:
            if data[k]:
                if isinstance(data[k], dict):
                    for pk, pv in data[k].items():
                        if pv:
                            result["xlsx_data"][k][pk] = pv
                elif data[k]:
                    result["xlsx_data"][k] = data[k]

    # Collect images (from folder and person subdirs)
    img_exts = {".png", ".jpg", ".jpeg"}
    for f in sorted(folder.rglob("*")):
        if f.suffix.lower() in img_exts:
            rel = str(f.relative_to(folder))
            person_hint = ""
            for p in ["BB", "LN"]:
                if f.parent.name == p or p in rel:
                    person_hint = p
                    break
            result["images"].append({
                "filename": rel,
                "path": str(f),
                "person_hint": person_hint,
                "size_bytes": f.stat().st_size,
            })

    # Summary
    bb_exp = len(result["xlsx_data"]["expenses"]["BB"])
    ln_exp = len(result["xlsx_data"]["expenses"]["LN"])
    bb_inc = len(result["xlsx_data"]["income"]["BB"])
    ln_inc = len(result["xlsx_data"]["income"]["LN"])
    img_count = len(result["images"])

    print(f"  → {bb_exp} BB expenses, {ln_exp} LN expenses", file=sys.stderr)
    print(f"  → {bb_inc} BB income, {ln_inc} LN income", file=sys.stderr)
    print(f"  → {img_count} images found", file=sys.stderr)

    return result


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Parse month folder for import")
    parser.add_argument("folder", help="Path to the month folder")
    parser.add_argument("--output", "-o", help="Output JSON file path")
    parser.add_argument("--stdout", action="store_true", help="Print JSON to stdout")
    args = parser.parse_args()

    result = scan_folder(Path(args.folder))

    output = json.dumps(result, ensure_ascii=False, indent=2, default=str)

    if args.stdout:
        print(output)
    elif args.output:
        Path(args.output).write_text(output, encoding="utf-8")
        print(f"Manifest written to {args.output}", file=sys.stderr)
    else:
        # Write to _manifest.json in the folder
        manifest_path = Path(args.folder) / "_manifest.json"
        manifest_path.write_text(output, encoding="utf-8")
        print(f"Manifest written to {manifest_path}", file=sys.stderr)


if __name__ == "__main__":
    main()
